from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlwt import Workbook
import requests
from bs4 import BeautifulSoup
import re
from selenium import webdriver
from xlwt import Workbook
from collections import Counter
from requests import get
import time

found_company = []
import pandas as pd

table = pd.read_excel('SCLog.xlsx',
                  sheet_name = 'SCLog',
                  header = 0,
                  usecols = "C",
                  convert_float = True)

browser = webdriver.Chrome()

for i in range(1,len(table)):
    browser.get(('https://www.sec.gov/edgar/searchedgar/companysearch.html'))
    find_company = browser.find_element_by_id('lesscompany')
    find_company.send_keys(table['Company Name1'][i])
    search_button = browser.find_element_by_name('Find').click()
    time.sleep(5)
    a = browser.page_source
    html_soup = BeautifulSoup(a, 'html.parser')
    source = html_soup.find_all('div', class_= 'noCompanyMatch')
    for i in range(1, len(source)):
        if source[i].text=="No matching companies.":
            browser.find_element_by_link_text('Company Search').click()
        else:
            found_company.append(table['Company Name1'][i])
            browser.find_element_by_link_text('Company Search').click()
            print(found_company)

#search_success = browser.find_elements_by_xpath('//*[@id="contentDiv"]/div')
#browser.find_elements_by_link_text('Search the Next-Generation EDGAR System')


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlwt import Workbook
import requests
from bs4 import BeautifulSoup
import re
from selenium import webdriver
from xlwt import Workbook
from collections import Counter

log_src = []
browser = webdriver.Chrome()
browser.get(('https://sclogisticsindustry.com/doing-business/'))
innerHTML = browser.execute_script("return document.body.innerHTML") 
html_soup = BeautifulSoup(innerHTML, 'html.parser')
source = html_soup.find_all('a', class_='place_title')
for i in range(1, len(source)):
    log_src.append(source[i].text)
import pandas as pd
test_df = pd.DataFrame({'Company Name':log_src})
test_df.to_csv("SCLog.csv")


#############################################FINAL CODE

#Extracting company names from SCLogistics website
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlwt import Workbook
import requests
from bs4 import BeautifulSoup
import re
from selenium import webdriver
from xlwt import Workbook
from collections import Counter

log_src = []
browser = webdriver.Chrome()
browser.get(('https://sclogisticsindustry.com/doing-business/'))
innerHTML = browser.execute_script("return document.body.innerHTML") 
html_soup = BeautifulSoup(innerHTML, 'html.parser')
html_soup
source = html_soup.find_all('a', class_='place_title')
for i in range(1, len(source)):
    log_src.append(source[i].text)
import pandas as pd
test_df = pd.DataFrame({'Company Name':log_src})
test_df.to_csv("SCLog_Results.csv")

#final code

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlwt import Workbook
import requests
from bs4 import BeautifulSoup
import re
from selenium import webdriver
from xlwt import Workbook
from collections import Counter
from requests import get
import time

found_company = []
page_found = []
not_found = []
import pandas as pd

table = pd.read_excel('SCLog.xlsx',
                  sheet_name = 'SCLog',
                  header = 0,
                  usecols = "D",
                  convert_float = True)

print("Number of null values are : {}".format(table['Names'].isnull().sum())) 

browser = webdriver.Chrome()

for i in range(0,len(table)):
    browser.get(('https://www.sec.gov/edgar/searchedgar/companysearch.html'))
    find_company = browser.find_element_by_id('lesscompany')
    find_company.send_keys(table['Names'][i])
    search_button = browser.find_element_by_name('Find').click()
    time.sleep(3)
    a = browser.page_source
    html_soup = BeautifulSoup(a, 'html.parser')
    source = html_soup.find_all('div', class_= 'noCompanyMatch')
    if len(source)==0:
        page_found = html_soup.find('span', class_='companyName')
        found_company.append(table['Names'][i])      
    else:
        for j in range(1, len(source)):
            not_found.append(table['Names'][i])
            if source[j].text=="No matching companies.":
                browser.find_element_by_link_text('Company Search').click()
import pandas as pd
test_df = pd.DataFrame({'Company Names':found_company})
#test_df1 = pd.DataFrame({'Companies Not Found:':not_found})
test_df.to_csv("Companies Found.csv")
#test_df1.to_csv("Companies Not Found.csv")

#final code with CIK updated
import csv
import xlsxwriter
import openpyxl
from openpyxl import load_workbook

write_Values = []
table = pd.read_excel('jsonfile.xlsx',
                  sheet_name = 'jsonfile',
                  header = 0,
                  usecols = "B",
                  convert_float = True)

browser = webdriver.Chrome()
for i in range(1,len(table)):
    browser.get(('https://www.sec.gov/edgar/searchedgar/companysearch.html'))
    find_company = browser.find_element_by_id('lesscompany')
    find_company.send_keys(table['Company Name'][i])
    search_button = browser.find_element_by_name('Find').click()
    time.sleep(3)
    a = browser.page_source
    html_soup = BeautifulSoup(a, 'html.parser')
    source = html_soup.find_all('div', class_= 'noCompanyMatch')
    if len(source)==0:
        page_found = html_soup.find('span', class_='companyName')     
        cik_no = re.findall("\d+", str(page_found))
        if not cik_no:
            print('CIK not found')
        else:
            myworkbook = openpyxl.load_workbook('jsonfile.xlsx')
            worksheet = myworkbook['jsonfile']
            worksheet.cell(row=i+2, column=3).value = cik_no[0]
            print(i)
            myworkbook.save('jsonfile.xlsx')
    else:
        for j in range(1, len(source)):
            if source[j].text=="No matching companies.":
                browser.find_element_by_link_text('Company Search').click()    
    
    
   #Recursive Search COde
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from xlwt import Workbook
import requests
from bs4 import BeautifulSoup
import numpy as np
import re
from nltk.corpus import stopwords
from selenium import webdriver
from xlwt import Workbook
from collections import Counter
from requests import get
import time

found_company = []
page_found = []
new_table = []
import pandas as pd

filtered_sentence = []
cik = []
company = []
region = []
table = pd.read_excel('rescanning.xlsx')
for i in range(0,len(table)):
    names = table['company_name'][i].lower()
    stop_words = set(stopwords.words('english'))
    word_tokens = word_tokenize(names) 
    tokens = [w for w in word_tokens if not w in stop_words]
    hu = ' '.join(tokens[:2])
    new_table.append(hu)
browser = webdriver.Chrome()
for i in range(0,len(new_table)):
    browser.get(('https://www.sec.gov/edgar/searchedgar/companysearch.html'))
    find_company = browser.find_element_by_id('lesscompany')
    find_company.send_keys(new_table[i].upper())
    search_button = browser.find_element_by_name('Find').click()
    time.sleep(3)
    a = browser.page_source
    html_soup = BeautifulSoup(a, 'html.parser')
    fulltable = html_soup.find('table', class_='tableFile2')
    for row in html_soup.find_all('tr')[2:]:
        tds = row.find_all('td')
        cik.append(tds[0].text)
        company.append(tds[1].text)
        region.append(tds[2].text)

import pandas as pd
recursive = pd.DataFrame({'CIK':cik, 'Company':company, 'Region':region})
recursive.to_csv("Recursive.csv")      
